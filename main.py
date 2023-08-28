#######################################################################################################################
from PyQt5.QtWidgets import QMessageBox
from PyQt5.uic import loadUi
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog
import os
########################################################################################################################
import openpyxl
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from PyQt5.QtCore import QTimer
#######################################################################################################################

class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        loadUi("CG_DATA.ui", self)
        self.setFixedSize(591, 207)
        self.pushButton.clicked.connect(self.cgdata)
        self.pushButton_3.clicked.connect(self.create)
        self.pushButton_2.clicked.connect(self.clear)
        self.radioButton.toggled.connect(self.on_radio_button_toggled)
        self.cgdata1 = None


    def cgdata(self):
        file_path1, _ = QFileDialog.getOpenFileName(self, "CG_DATA file", os.path.expanduser("~"), "CG_DATA (*.txt)")
        self.lineEdit.setText(file_path1)
        self.cgdata1 = file_path1

    def create(self):
        aranan_degerler = [
            "SIP-Method",
        ]

        aranan_degerler2 = ["Time-Stamps"]

        aranan_degerler3 = [
            " Session-Id",
        ]

        aranan_degerler4 = [
            "User-Session-Id",
            "Node-Functionality",
            "Original-Caller-Party-Number",
            "Calling-Party-Address",
            "Called-Party-Address",
            "IMS-Charging-Identifier",
            "Dialled-Party-Address",
            "Outgoing-Trunk-Group-Id",
            "Incoming-Trunk-Group-Id",
        ]
        output_list = []
        output_values = {}

        with open(self.cgdata1) as f1:
            lines = f1.readlines()
            for i, satir in enumerate(lines):
                for j, aranan_deger in enumerate(aranan_degerler):
                    if aranan_deger in satir:
                        for k in range(i + 5, i + 6):
                            if k < len(lines):
                                output = lines[k].strip().replace(" ", "").replace("AVPData(UTF8String):", "")
                                output_values[aranan_deger] = output

                for j, aranan_deger in enumerate(aranan_degerler2):
                    if aranan_deger in satir:
                        for k in range(i + 11, i + 12):
                            if k < len(lines):
                                output = lines[k].strip().replace(" ", "")
                                if output.startswith("AVPData(Time):"):
                                    output = output.replace("AVPData(Time):", "")[10:19]
                                    saat_dt = datetime.strptime(output, "%H:%M:%S")
                                    saat_dt = saat_dt + timedelta(hours=3)
                                    yeni_saat = saat_dt.strftime("%H:%M:%S")
                                    output_values[aranan_deger] = yeni_saat

                for j, aranan_deger in enumerate(aranan_degerler3):
                    if aranan_deger in satir:
                        for k in range(i + 4, i + 5):
                            if k < len(lines):
                                output = lines[k].strip().replace(" ", "").replace("AVPData(UTF8String):", "").replace(
                                    "AVPData(Enumerated):", "")
                                output_values[aranan_deger] = output

                for j, aranan_deger in enumerate(aranan_degerler4):
                    if aranan_deger in satir:
                        for k in range(i + 5, i + 6):
                            if k < len(lines):
                                output = lines[k].strip().replace(" ", "").replace("AVPData(UTF8String):", "").replace(
                                    "AVPData(Enumerated):", "")
                                output_values[aranan_deger] = output

                if satir.startswith("[Diameter Head]"):
                    if output_values:
                        output_list.append(output_values)
                    output_values = {}

            if output_values:
                output_list.append(output_values)

        # Çıktıları xlsx dosyasına kaydetme
        wb = openpyxl.Workbook()
        ws = wb.active

        # Başlık satırını yazma
        header_row = 1
        for col_num, header in enumerate(aranan_degerler + aranan_degerler2 + aranan_degerler3 + aranan_degerler4,start=1):
            ws.cell(row=header_row, column=col_num).value = header

        # Çıktıları yazma
        data_row = 2
        for output_values in output_list:
            col_num = 1
            for aranan_deger in aranan_degerler + aranan_degerler2 + aranan_degerler3 + aranan_degerler4:
                output = output_values.get(aranan_deger, "")
                ws.cell(row=data_row, column=col_num).value = output
                col_num += 1
            data_row += 1
        # Dosyayı kaydetme
        dosya_adi = "output.xlsx"
        wb.save(dosya_adi)
        QMessageBox.information(self, "Successful", "The output was created in the application directory")

    def on_radio_button_toggled(self):
        QTimer.singleShot(0, self.process_radio_button1)

    def process_radio_button1(self):
            def find_matching_value(filename, sheetname, column_index, value_column_index):
                workbook = load_workbook(filename)
                sheet = workbook[sheetname]
                column = sheet[column_index]
                value_column = sheet[value_column_index]
                matching_pairs = []
                # Duration sütununun başlığını ekleyin
                time_difference_column = sheet.max_column + 1
                time_difference_column_letter = get_column_letter(time_difference_column)
                sheet[f"{time_difference_column_letter}1"] = "Duration"

                for i in range(1, len(column)):
                    for j in range(i + 1, len(column)):
                        if column[i].value == column[j].value:
                            first_value = value_column[i].value
                            second_value = value_column[j].value
                            if first_value is not None and second_value is not None:
                                first_value = datetime.strptime(first_value, "%H:%M:%S")
                                second_value = datetime.strptime(second_value, "%H:%M:%S")
                                time_difference = second_value - first_value
                                pair = (column[i].value, first_value, second_value, time_difference)
                                matching_pairs.append(pair)
                                time_difference_formatted = str(timedelta(hours=time_difference.total_seconds() / 3600))
                                sheet[f"{time_difference_column_letter}{i + 1}"] = time_difference_formatted

                workbook.save(filename)

                return matching_pairs

            # Örnek kullanım
            filename = 'output.xlsx'
            sheetname = 'Sheet'
            column_index = 'C'
            value_column_index = 'B'
            find_matching_value(filename, sheetname, column_index, value_column_index)
            QMessageBox.information(self, "Successful", "The call time between users was calculated.")

    def clear(self):
        self.lineEdit.clear()




if __name__ == "__main__":
    app = QApplication(sys.argv)
    mainWindow = MainWindow()
    mainWindow.show()
    sys.exit(app.exec_())